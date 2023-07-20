import openpyxl

book = openpyxl.load_workbook("/home/inchara/Documents/cred_details.xlsx")
sheet = book.active
cell = sheet.cell(row=2, column=1)
print(cell.value)
for i in range(1, sheet.max_row+1):
    print(sheet.cell(row=i, column=1).value)


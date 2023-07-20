import logging
import time

import openpyxl
from selenium.webdriver.common.by import By

from PageObjects.login import Login
from Utilities import CustomLogger
from Utilities.readProperties import ReadConfig


class Testlogin:
    # loginpage = None
    logger = CustomLogger.setup_logger('Login', ReadConfig.get_logs_directory() + "/Test_login.log",
                                       level=logging.DEBUG)

    # book = openpyxl.load_workbook("/home/inchara/Documents/cred_details.xlsx")
    # sheet = book.active
    # username = sheet.cell(row=2, column=1)
    # password1 = sheet.cell(row=2, column=2)
    # user = username.value
    # print(user)
    # password = password1.value
    # print(password)

    # def test_login_page_tittle(self, setup):
    #     self.logger.info("***********************Test_001_Login*****************************")
    #     self.logger.info("********************test_login_page_tittle started*********************")
    #     self.driver = setup
    #     self.loginpage = Login(self.driver)
    #     self.loginpage.open_application()
    #     time.sleep(5)
    #     self.driver.find_element(By.NAME, "username").send_keys(self.user)
    #     self.driver.find_element(By.NAME, "password").send_keys(self.password)
    #     self.driver.find_element(By.XPATH, "//button[@type='submit']").click()
    #     time.sleep(4)

    def test_login3_page_tittle(self, setup):
        self.logger.info("***********************Test_001_Login*****************************")
        self.logger.info("********************test_login_page_tittle started*********************")
        self.driver = setup
        self.loginpage = Login(self.driver)
        self.loginpage.open_application()
        time.sleep(3)
        self.loginpage.open_logipage_xl()
